from openpyxl import load_workbook
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import shutil
import datetime
import glob
import os
import csv
from selenium.webdriver.chrome.options import Options
import msvcrt


wb = load_workbook(filename = 'Input.xlsx')
ws = wb.active
headerRowNo = ws["E5"].value
reportLink = ws["E9"].value
email = ws['E11'].value
password = ws['E13'].value
reportName = ws['I12'].value
reportName = str(reportName).split(" 0")[0]
repeatTime = ws['E7'].value
saveLocation = ws["I10"].value
saveLocation = saveLocation.replace("\\",'/')
downloads_folder = ws['E21'].value
downloads_folder = downloads_folder.replace("\\",'/')
chrome_driver = ws["E23"].value
chrome_driver = chrome_driver.replace("\\",'/')
refreshBTN = ws['J19'].value
csvBTN = ws['J21'].value
# print(chrome_driver)

fromDate = ws['J15'].value
fromDate = str(fromDate).split(" 0")[0]
toDate = ws['J17'].value
toDate = str(toDate).split(" 0")[0]

fromDate = datetime.datetime.strptime(fromDate, "%Y-%m-%d").strftime("%d/%m/%Y")
toDate = datetime.datetime.strptime(toDate, "%Y-%m-%d").strftime("%d/%m/%Y")


def main():
    try:
        chrome_options = Options()
        # chrome_options.add_argument("--headless")
        chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        driver = webdriver.Chrome(executable_path=chrome_driver, options=chrome_options)

        print("Logging In.....")
        driver.get(reportLink)
        driver.find_element_by_xpath("//input[@id='userName']").send_keys(email)
        driver.find_element_by_xpath("//input[@id='password']").send_keys(password)
        driver.find_element_by_xpath("//button[@id='login-submit']").click()
        time.sleep(5)
        ques = driver.find_element_by_xpath(
            "/html/body/div[2]/div[2]/form/table/tbody/tr[3]/td/table/tbody/tr[2]/td[2]").text
        ques_ans = driver.find_element_by_xpath("//input[@id='null']")
        ques_submit = driver.find_element_by_xpath("//input[@class='bgbutton']")

        questions = {
            ws['E15'].value: ws["G15"].value,
            ws['E17'].value: ws["G17"].value,
            ws['E19'].value: ws["G19"].value,

        }

        for key in questions:
            if key in ques:
                ques_ans.send_keys(questions[key])
                ques_submit.click()

        time.sleep(5)

        print("Logged In Successfully")

        print("Getting Data from " + fromDate + " to " + toDate + ".....")
        driver.find_element_by_xpath("//input[@id='crit_1_from']").clear()
        driver.find_element_by_xpath("//input[@id='crit_1_from']").send_keys(fromDate)

        driver.find_element_by_xpath("//input[@id='crit_1_to']").clear()
        driver.find_element_by_xpath("//input[@id='crit_1_to']").send_keys(toDate)

        driver.find_element_by_xpath(refreshBTN).click()

        try:
            wait = WebDriverWait(driver, 60)
            element = wait.until(EC.element_to_be_clickable(
                (By.XPATH, csvBTN))).click()

            print("Downloading File...")

        except:
            print("No Data Available on this date Range")
            pass

        def every_downloads_chrome(driver):
            if not driver.current_url.startswith("chrome://downloads"):
                driver.get("chrome://downloads/")
            return driver.execute_script("""
                var items = document.querySelector('downloads-manager')
                    .shadowRoot.getElementById('downloadsList').items;
                if (items.every(e => e.state === "COMPLETE"))
                    return items.map(e => e.fileUrl || e.file_url);
                """)

        # waits for all the files to be completed and returns the paths
        paths = WebDriverWait(driver, 120, 1).until(every_downloads_chrome)
        print(paths)

        print("File Downloading Completes")
        list_of_files = glob.glob(downloads_folder + "/*.csv")
        latest_file = max(list_of_files, key=os.path.getctime)


        input_file = open(latest_file, 'r+')
        output_file = open(saveLocation + "/" + reportName + ".csv", 'a+', newline='', encoding='utf-8-sig')
        writer = csv.writer(output_file)
        main_list = []

        for i in csv.reader(input_file):
            main_list.append(i)

        main_list = main_list[headerRowNo - 1:]
        for i in main_list:
            writer.writerow(i)

        input_file.close()
        output_file.close()

        driver.close()

        time.sleep(5)

        os.remove(latest_file)

    except Exception as e:
        print(e)
        driver.close()

while True:
    main()
    print("For EXIT Press (ESC)")
    time.sleep(5)

    if msvcrt.kbhit():
        if ord(msvcrt.getch()) == 27:
            break
    time.sleep(repeatTime * 60)

